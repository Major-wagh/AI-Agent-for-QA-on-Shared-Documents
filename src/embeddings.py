import os
import polars as pl
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_google_genai import GoogleGenerativeAIEmbeddings
from langchain_community.vectorstores import FAISS
from langchain.schema import Document
from dotenv import load_dotenv
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import pickle
import hashlib
import json
from pathlib import Path
load_dotenv()

class DocumentEmbeddings:
    def __init__(self):
        self.embeddings = None
        self.cache_dir = Path("./embedding_cache")
        self.cache_dir.mkdir(exist_ok=True)
    
    def get_file_hash(self, file_path):
        """Generate hash for file to check if it's changed"""
        file_path = Path(file_path)
        stat = file_path.stat()
        # Use file size and modification time for faster hashing
        hash_input = f"{file_path.name}_{stat.st_size}_{stat.st_mtime}"
        return hashlib.md5(hash_input.encode()).hexdigest()
    
    def load_excel_with_polars(self, file_path, sheet_name=None):
        """Load Excel file using Polars - much faster than pandas"""
        try:
            print(f"Loading {file_path} with Polars...")
            file_path = Path(file_path)
            
            # Get all sheet names first
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            documents = []
            
            for sheet in sheet_names:
                try:
                    print(f"  Processing sheet: {sheet}")
                    
                    # Read with Polars - much faster than pandas
                    df = pl.read_excel(
                        file_path,
                        sheet_name=sheet,
                        read_csv_options={
                            "ignore_errors": True,
                            "null_values": ["", "NULL", "null", "N/A", "n/a"]
                        }
                    )
                    
                    # Fast filtering with Polars
                    if df.height == 0:
                        continue
                    
                    # Remove completely empty columns efficiently
                    non_empty_cols = []
                    for col in df.columns:
                        if not df[col].is_null().all():
                            non_empty_cols.append(col)
                    
                    if not non_empty_cols:
                        continue
                    
                    df_clean = df.select(non_empty_cols)
                    
                    # Remove completely empty rows efficiently
                    df_clean = df_clean.filter(
                        ~pl.all_horizontal(pl.col("*").is_null())
                    )
                    
                    if df_clean.height == 0:
                        continue
                    
                    # Convert to string efficiently with Polars
                    text_content = self.polars_df_to_text(df_clean, sheet)
                    
                    # Create document
                    doc = Document(
                        page_content=text_content,
                        metadata={
                            "source": str(file_path),
                            "sheet": sheet,
                            "rows": df_clean.height,
                            "columns": df_clean.width,
                            "file_size": file_path.stat().st_size,
                            "file_type": "xlsx"
                        }
                    )
                    documents.append(doc)
                    
                except Exception as e:
                    print(f"    Error processing sheet {sheet}: {e}")
                    continue
            
            print(f"Loaded {len(documents)} sheets from {file_path}")
            return documents
            
        except Exception as e:
            print(f"Error loading {file_path} with Polars: {e}")
            return []
    
    def polars_df_to_text(self, df, sheet_name):
        """Convert Polars DataFrame to text efficiently"""
        # Get column info
        columns = df.columns
        dtypes = [str(dtype) for dtype in df.dtypes]
        
        text_parts = [
            f"Sheet: {sheet_name}",
            f"Dimensions: {df.height} rows × {df.width} columns",
            f"Columns: {', '.join(columns)}",
            f"Data Types: {dict(zip(columns, dtypes))}",
            "",  # Empty line separator
        ]
        
        # Sample data efficiently - get first 100 rows or all if less
        sample_size = min(100, df.height)
        sample_df = df.head(sample_size)
        
        # Convert to string representation
        # For very wide tables, limit columns shown
        if df.width > 20:
            # Show first 10 and last 10 columns
            display_df = pl.concat([
                sample_df.select(columns[:10]),
                sample_df.select(columns[-10:])
            ], how="horizontal")
            text_parts.append("Sample data (showing first 10 and last 10 columns):")
        else:
            display_df = sample_df
            text_parts.append("Sample data:")
        
        # Convert to string efficiently
        text_parts.append(str(display_df))
        
        # Add summary statistics for numeric columns
        numeric_cols = df.select(pl.col(pl.NUMERIC_DTYPES)).columns
        if numeric_cols:
            text_parts.append("\nNumeric column summaries:")
            for col in numeric_cols[:5]:  # Limit to first 5 numeric columns
                try:
                    stats = df[col].describe()
                    text_parts.append(f"{col}: {stats}")
                except:
                    continue
        
        return "\n".join(text_parts)
    
    def load_document(self, file_path):
        """Load document based on file extension"""
        file_extension = Path(file_path).suffix.lower()
        if file_extension == ".xlsx":
            return self.load_excel_with_polars(file_path)
        else:
            print(f"Unsupported file type: {file_extension} for {file_path}")
            return []
    
    def process_file_with_cache(self, file_path, use_streaming=False, chunk_rows=50000):
        """Process single file with caching"""
        file_hash = self.get_file_hash(file_path)
        cache_file = self.cache_dir / f"{Path(file_path).stem}_{file_hash}.pkl"
        
        # Check if cached version exists
        if cache_file.exists():
            print(f"Loading cached chunks for {file_path}")
            try:
                with open(cache_file, 'rb') as f:
                    return pickle.load(f)
            except Exception as e:
                print(f"Cache corrupted for {file_path}, reprocessing...")
        
        # Process file
        documents = self.load_document(file_path) # Use the generic loader
        
        if not documents:
            return []
        
        # Chunk documents
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            length_function=len,
            add_start_index=True,
        )
        chunks = text_splitter.split_documents(documents)
        
        # Cache the chunks
        try:
            with open(cache_file, 'wb') as f:
                pickle.dump(chunks, f)
            print(f"Processed and cached {len(chunks)} chunks from {file_path}")
        except Exception as e:
            print(f"Warning: Could not cache chunks for {file_path}: {e}")
        
        return chunks
    
    def create_embeddings_parallel(self, all_chunks, batch_size=100, max_workers=3):
        """Create embeddings with parallel processing"""
        if not self.embeddings:
            self.embeddings = GoogleGenerativeAIEmbeddings(
                model="models/text-embedding-004",
                task_type="semantic_similarity"
            )
        
        print(f"Creating embeddings for {len(all_chunks)} chunks using parallel processing")
        
        # Split chunks into batches
        batches = [all_chunks[i:i + batch_size] for i in range(0, len(all_chunks), batch_size)]
        
        def process_batch(batch_data):
            batch_idx, batch_chunks = batch_data
            try:
                print(f"Processing batch {batch_idx + 1}/{len(batches)}")
                batch_texts = [doc.page_content for doc in batch_chunks]
                
                # Add delay to avoid rate limiting
                if batch_idx > 0:
                    time.sleep(0.5)
                
                # Create embeddings for this batch
                vectors = self.embeddings.embed_documents(batch_texts)
                return batch_chunks, vectors
                
            except Exception as e:
                print(f"Error processing batch {batch_idx + 1}: {e}")
                return batch_chunks, None
        
        # Process batches
        processed_chunks = []
        all_vectors = []
        
        # Use ThreadPoolExecutor for I/O bound embedding calls
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(process_batch, (i, batch)): i 
                for i, batch in enumerate(batches)
            }
            
            for future in as_completed(futures):
                try:
                    chunks, vectors = future.result()
                    if vectors is not None:
                        processed_chunks.extend(chunks)
                        all_vectors.extend(vectors)
                except Exception as e:
                    print(f"Error in parallel processing: {e}")
        
        return processed_chunks, all_vectors
    
    def create_and_store_embeddings_optimized(self, files_to_embed, use_streaming=False, 
                                            chunk_rows=50000, use_parallel=True):
        """
        Ultra-optimized version with Polars, caching, streaming, and parallel processing
        Handles multiple file types.
        """
        if not os.getenv("GOOGLE_API_KEY"):
            print("Error: GOOGLE_API_KEY environment variable not set!")
            return
        
        start_time = time.time()
        all_chunks = []
        
        print(f"Processing {len(files_to_embed)} files with optimized loaders...")
        
        # Process files
        for file_path in files_to_embed:
            if not Path(file_path).exists():
                print(f"Warning: {file_path} not found. Skipping...")
                continue
            
            file_size = Path(file_path).stat().st_size / (1024 * 1024)  # MB
            print(f"File size: {file_size:.1f} MB")
            
            # Auto-decide streaming based on file size for Excel files
            use_file_streaming = use_streaming or (file_size > 50 and Path(file_path).suffix.lower() == ".xlsx")
            
            chunks = self.process_file_with_cache(
                file_path, 
                use_streaming=use_file_streaming,
                chunk_rows=chunk_rows
            )
            all_chunks.extend(chunks)
        
        if not all_chunks:
            print("No chunks created. Exiting...")
            return
        
        print(f"Total chunks: {len(all_chunks)}")
        
        try:
            # Initialize embeddings
            if not self.embeddings:
                self.embeddings = GoogleGenerativeAIEmbeddings(
                    model="models/text-embedding-004",
                    task_type="semantic_similarity"
                )
            
            # Create vector store
            if use_parallel and len(all_chunks) > 100:
                print("Using parallel embedding creation...")
                processed_chunks, vectors = self.create_embeddings_parallel(all_chunks)
                
                # Create FAISS index from embeddings
                vector_store = FAISS.from_documents(processed_chunks, self.embeddings)
            else:
                print("Using sequential embedding creation...")
                vector_store = FAISS.from_documents(all_chunks, self.embeddings)
            
            # Save vector store
            faiss_index_path = "faiss_excel_index"
            vector_store.save_local(faiss_index_path)
            
            end_time = time.time()
            print(f"Process completed in {end_time - start_time:.2f} seconds!")
            print(f"FAISS index saved at {faiss_index_path}")
            print(f"Speed: {len(all_chunks) / (end_time - start_time):.1f} chunks/second")
            
            return vector_store
            
        except Exception as e:
            print(f"Error creating vector store: {e}")
            return None
    
    def clear_cache(self):
        """Clear all cached files"""
        import shutil
        if self.cache_dir.exists():
            shutil.rmtree(self.cache_dir)
            self.cache_dir.mkdir()
            print("Cache cleared")

def create_and_store_embeddings(files_to_embed: list = None):
    """Main function using Polars optimization"""
    if files_to_embed is None:
        files_to_embed = ["./data/data.xlsx", "./data/Forcast.xlsx"]
    
    processor = DocumentEmbeddings() # Changed class name
    return processor.create_and_store_embeddings_optimized(
        files_to_embed, 
        use_streaming=True,  # Auto-enables for large Excel files
        chunk_rows=50000,    # Rows per chunk for streaming
        use_parallel=True    # Parallel embedding creation
    )

def load_existing_vector_store():
    """Load existing FAISS vector store"""
    try:
        embeddings = GoogleGenerativeAIEmbeddings(
            model="models/text-embedding-004",
            task_type="semantic_similarity"
        )
        vector_store = FAISS.load_local(
            "faiss_excel_index", 
            embeddings, 
            allow_dangerous_deserialization=True
        )
        return vector_store
    except Exception as e:
        print(f"Error loading existing vector store: {e}")
        return None

# Utility functions
def analyze_excel_file(file_path):
    """Quick analysis of Excel file structure with Polars"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        print(f"\nFile: {file_path}")
        print(f"File size: {Path(file_path).stat().st_size / (1024*1024):.1f} MB")
        print(f"Sheets: {len(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            try:
                df = pl.read_excel(file_path, sheet_name=sheet_name)
                print(f"  {sheet_name}: {df.height:,} rows × {df.width} columns")
                
                # Show data types
                numeric_cols = len(df.select(pl.col(pl.NUMERIC_DTYPES)).columns)
                text_cols = df.width - numeric_cols
                print(f"    Numeric columns: {numeric_cols}, Text columns: {text_cols}")
                
            except Exception as e:
                print(f"  {sheet_name}: Error reading - {e}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing {file_path}: {e}")

def benchmark_methods(file_path, iterations=3):
    """Benchmark Polars vs Pandas for Excel loading"""
    import pandas as pd
    
    print(f"Benchmarking Excel loading methods for {file_path}")
    
    # Polars timing
    polars_times = []
    for i in range(iterations):
        start = time.time()
        try:
            df = pl.read_excel(file_path)
            polars_times.append(time.time() - start)
        except Exception as e:
            print(f"Polars error: {e}")
            break
    
    # Pandas timing
    pandas_times = []
    for i in range(iterations):
        start = time.time()
        try:
            df = pd.read_excel(file_path)
            pandas_times.append(time.time() - start)
        except Exception as e:
            print(f"Pandas error: {e}")
            break
    
    if polars_times and pandas_times:
        avg_polars = sum(polars_times) / len(polars_times)
        avg_pandas = sum(pandas_times) / len(pandas_times)
        speedup = avg_pandas / avg_polars
        
        print(f"Average Polars time: {avg_polars:.2f}s")
        print(f"Average Pandas time: {avg_pandas:.2f}s")
        print(f"Polars speedup: {speedup:.1f}x faster")

if __name__ == "__main__":
    create_and_store_embeddings()
